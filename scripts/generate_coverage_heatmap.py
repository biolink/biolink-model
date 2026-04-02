#!/usr/bin/env python3
"""
Generate an interactive HTML heatmap showing ingest coverage across the biolink entity matrix.

Reads the filled-in biolink_entity_matrix.xlsx spreadsheet and produces an HTML file where:
- Cell color intensity shows the number of ingests covering each subject->object entity pair
- Hovering reveals the list of ingests per cell
- Only entity classes with at least one ingest are shown

Usage:
    uv run python scripts/generate_coverage_heatmap.py [--input PATH] [--output PATH]
"""

import argparse
import openpyxl


def generate_heatmap(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb['Entity Matrix - Data Sources']

    # Build headers
    headers_row = []
    for c in range(2, ws.max_column + 1):
        headers_row.append(ws.cell(row=1, column=c).value)

    row_labels = []
    for r in range(2, ws.max_row + 1):
        row_labels.append(ws.cell(row=r, column=1).value)

    # Collect ingest coverage data
    matrix = []
    for r_idx in range(len(row_labels)):
        r = r_idx + 2
        row_data = []
        for c_idx in range(len(headers_row)):
            c = c_idx + 2
            ingests = ws.cell(row=r, column=c).value
            ingest_list = [x.strip() for x in ingests.split(',')] if ingests else []
            row_data.append({
                'ingests': ingests or '',
                'count': len(ingest_list),
            })
        matrix.append(row_data)

    # Filter to entity classes with at least one filled cell
    active_rows = set()
    active_cols = set()
    for r_idx in range(len(row_labels)):
        for c_idx in range(len(headers_row)):
            if matrix[r_idx][c_idx]['count'] > 0:
                active_rows.add(r_idx)
                active_cols.add(c_idx)

    active_rows = sorted(active_rows)
    active_cols = sorted(active_cols)

    max_count = max(
        (matrix[r][c]['count'] for r in active_rows for c in active_cols),
        default=1
    )
    filled_count = sum(1 for r in active_rows for c in active_cols if matrix[r][c]['count'] > 0)

    print(f"Active: {len(active_rows)} rows x {len(active_cols)} cols")
    print(f"Max ingests per cell: {max_count}")
    print(f"Filled cells: {filled_count}")

    def cell_color(count):
        """Map ingest count to an RGB color string."""
        if count == 0:
            return '#16213e'
        frac = count / max_count
        if frac < 0.5:
            r = int(20 + frac * 2 * 20)
            g = int(40 + frac * 2 * 180)
            b = int(80 + frac * 2 * 40)
        else:
            f2 = (frac - 0.5) * 2
            r = int(40 + f2 * 215)
            g = int(220 - f2 * 30)
            b = int(120 - f2 * 100)
        return f'rgb({r},{g},{b})'

    # Generate HTML
    html_parts = ["""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Biolink Entity Matrix - Ingest Coverage</title>
<style>
body { font-family: -apple-system, BlinkMacSystemFont, sans-serif; margin: 20px; background: #1a1a2e; color: #eee; }
h1 { color: #e94560; }
.legend { display: flex; gap: 4px; margin: 15px 0; align-items: center; }
.legend-label { font-size: 13px; color: #aaa; margin: 0 8px; }
.legend-swatch { width: 18px; height: 18px; border: 1px solid #333; border-radius: 2px; }
.stats { margin: 15px 0; font-size: 14px; color: #aaa; }
.matrix-container { overflow: auto; max-height: 85vh; max-width: 95vw; }
table { border-collapse: collapse; font-size: 10px; }
th { position: sticky; background: #16213e; z-index: 2; padding: 2px 4px; border: 1px solid #333; font-weight: normal; }
th.row-header { left: 0; z-index: 3; text-align: right; white-space: nowrap; min-width: 120px; }
th.col-header { top: 0; z-index: 2; white-space: nowrap; min-width: 22px; max-width: 22px; height: 160px; vertical-align: bottom; }
th.corner { left: 0; top: 0; z-index: 4; }
td { width: 22px; height: 22px; min-width: 22px; border: 1px solid #222; padding: 0; cursor: default; }
td:hover { outline: 2px solid #fff; outline-offset: -2px; z-index: 1; position: relative; }
#tooltip { display: none; position: fixed; background: #0f3460; border: 1px solid #e94560;
  padding: 10px; border-radius: 6px; max-width: 500px; font-size: 12px; z-index: 100;
  box-shadow: 0 4px 12px rgba(0,0,0,0.5); pointer-events: none; }
#tooltip .tt-title { font-weight: bold; color: #e94560; margin-bottom: 6px; }
#tooltip .tt-label { color: #aaa; }
</style>
</head>
<body>
<h1>Biolink Entity Matrix &mdash; Ingest Coverage Heatmap</h1>
<p style="color:#aaa">Cell intensity shows number of ingests covering each subject&rarr;object entity pair. Hover for details.</p>

<div class="legend">
  <span class="legend-label">0</span>
"""]

    # Color scale legend
    for i in range(max_count + 1):
        if i % max(1, max_count // 10) == 0 or i == max_count:
            html_parts.append(
                f'  <div class="legend-swatch" style="background:{cell_color(i)}"></div>\n'
            )

    html_parts.append(f"""  <span class="legend-label">{max_count}</span>
</div>
<div class="stats">
  Showing {len(active_rows)} x {len(active_cols)} entity classes with at least one ingest &nbsp;|&nbsp;
  Filled cells: <b>{filled_count}</b>
</div>

<div id="tooltip"></div>

<div class="matrix-container">
<table>
<tr><th class="corner"></th>
""")

    # Column headers
    for c_idx in active_cols:
        label = headers_row[c_idx] or ''
        html_parts.append(
            f'<th class="col-header"><div style="transform:rotate(-60deg);'
            f'transform-origin:left bottom;white-space:nowrap;width:20px">'
            f'{label}</div></th>\n'
        )
    html_parts.append('</tr>\n')

    # Data rows
    for r_idx in active_rows:
        label = row_labels[r_idx] or ''
        html_parts.append(f'<tr><th class="row-header">{label}</th>\n')
        for c_idx in active_cols:
            cell = matrix[r_idx][c_idx]
            bg = cell_color(cell['count'])
            subj = row_labels[r_idx]
            obj = headers_row[c_idx]
            ingests_esc = cell['ingests'].replace("'", "\\'")

            html_parts.append(
                f'<td style="background:{bg}" '
                f"onmouseover=\"showTip(event,'{subj}','{obj}',{cell['count']},'{ingests_esc}')\" "
                f'onmouseout="hideTip()"></td>\n'
            )
        html_parts.append('</tr>\n')

    html_parts.append("""</table>
</div>

<script>
const tip = document.getElementById('tooltip');
function showTip(e, subj, obj, count, ingests) {
  tip.innerHTML = `<div class="tt-title">${subj} &rarr; ${obj}</div>
    <div><span class="tt-label">Ingests (${count}):</span><br>${ingests || 'none'}</div>`;
  tip.style.display = 'block';
  tip.style.left = (e.clientX + 15) + 'px';
  tip.style.top = (e.clientY + 15) + 'px';
}
function hideTip() { tip.style.display = 'none'; }
</script>
</body>
</html>
""")

    with open(output_path, 'w') as f:
        f.write(''.join(html_parts))
    print(f"Wrote {output_path}")


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('--input', type=str, default='biolink_entity_matrix.xlsx',
                        help='Input xlsx file (default: biolink_entity_matrix.xlsx)')
    parser.add_argument('--output', type=str, default='src/docs/coverage_heatmap.html',
                        help='Output HTML file (default: src/docs/coverage_heatmap.html)')
    args = parser.parse_args()

    generate_heatmap(args.input, args.output)


if __name__ == '__main__':
    main()
