#!/usr/bin/env python3
"""
Generate the biolink entity matrix spreadsheet with ingest coverage data and example edge comments.

This script:
1. Creates a 157x157 matrix of biolink entity classes (all named thing descendants, excluding associations)
2. Fetches graph-metadata.json for each ingest from the latest release at kgx-storage.rtx.ai
3. For each ingest edge type, propagates UPWARD through the biolink is_a and mixin hierarchy
4. Fills in cells with the names of ingests that cover each subject-object entity pair
5. Scans local edge data (if available) to add one example edge per ingest per cell as xlsx comments

Usage:
    uv run python scripts/generate_entity_matrix.py [--local-data PATH]

Arguments:
    --local-data PATH   Path to translator-ingests/data/ directory for edge examples.
                        If omitted, only ingest names are filled in (no comments).
"""

import argparse
import json
import re
import urllib.request
import ssl
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, PatternFill, Font
from linkml_runtime.utils.schemaview import SchemaView

RELEASE_SUMMARY_URL = "https://kgx-storage.rtx.ai/releases/latest-release-summary.json"
MAX_EDGES_SCAN = 500_000  # edges to scan per ingest per pass


def curie_to_class_name(curie):
    """Convert biolink:CamelCase to linkml space-separated lowercase name."""
    name = curie.replace("biolink:", "")
    name = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', name)
    name = re.sub(r'(?<=[A-Z])(?=[A-Z][a-z])', ' ', name)
    return name.lower()


def build_hierarchy(sv):
    """Build hierarchy lookup structures from the biolink schema."""
    all_classes = set(sv.all_classes())
    assoc_descendants = set(sv.class_descendants('association'))
    named_thing_descendants = set(sv.class_descendants('named thing'))
    entity_classes = named_thing_descendants - assoc_descendants

    # Precompute: for each biolink class, which entity classes does it map to (upward only)?
    category_to_entity_classes = {}
    for cls in all_classes:
        headers = set()
        for anc in sv.class_ancestors(cls, mixins=True):
            if anc in entity_classes:
                headers.add(anc)
        if headers:
            category_to_entity_classes[cls] = headers

    return entity_classes, category_to_entity_classes


def get_entity_ancestors(class_name, category_to_entity_classes):
    """Get entity class ancestors for a class (upward propagation only)."""
    return category_to_entity_classes.get(class_name, set())


def create_blank_matrix(entity_classes, sv):
    """Create a blank xlsx with entity class headers on both axes."""
    wb = openpyxl.Workbook()

    # Sort entity classes alphabetically
    sorted_classes = sorted(entity_classes)

    # ── Tab 1: Entity Matrix - Data Sources ──
    ws1 = wb.active
    ws1.title = "Entity Matrix - Data Sources"

    # Write headers
    for i, cls in enumerate(sorted_classes):
        ws1.cell(row=1, column=i + 2, value=cls)
        ws1.cell(row=i + 2, column=1, value=cls)
        # Horizontal text for column headers
        ws1.cell(row=1, column=i + 2).alignment = Alignment(text_rotation=0)

    # ── Tab 2: Association Constraints ──
    ws2 = wb.create_sheet("Association Constraints")

    for i, cls in enumerate(sorted_classes):
        ws2.cell(row=1, column=i + 2, value=cls)
        ws2.cell(row=i + 2, column=1, value=cls)
        ws2.cell(row=1, column=i + 2).alignment = Alignment(text_rotation=0)

    # Fill in association constraints
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    row_idx = {cls: i + 2 for i, cls in enumerate(sorted_classes)}
    col_idx = {cls: i + 2 for i, cls in enumerate(sorted_classes)}
    entity_set = set(sorted_classes)

    for assoc_name in sv.class_descendants('association'):
        assoc_cls = sv.get_class(assoc_name)
        if assoc_cls is None:
            continue

        induced_subj = sv.get_slot('subject', cls=assoc_cls)
        induced_obj = sv.get_slot('object', cls=assoc_cls)

        subj_range = induced_subj.range if induced_subj else None
        obj_range = induced_obj.range if induced_obj else None

        if not subj_range or not obj_range:
            continue

        subj_classes = set(sv.class_descendants(subj_range, mixins=True)) & entity_set
        obj_classes = set(sv.class_descendants(obj_range, mixins=True)) & entity_set

        for s in subj_classes:
            for o in obj_classes:
                r, c = row_idx[s], col_idx[o]
                existing = ws2.cell(row=r, column=c).value
                if existing:
                    ws2.cell(row=r, column=c).value = existing + '\n' + assoc_name
                else:
                    ws2.cell(row=r, column=c).value = assoc_name
                    ws2.cell(row=r, column=c).fill = green_fill

    return wb


def fetch_release_summary():
    """Fetch the latest release summary from kgx-storage."""
    ctx = ssl.create_default_context()
    with urllib.request.urlopen(RELEASE_SUMMARY_URL, context=ctx) as resp:
        return json.loads(resp.read())


def fill_ingest_data(wb, category_to_entity_classes):
    """Fetch graph-metadata.json for each ingest and fill in the matrix."""
    ws = wb["Entity Matrix - Data Sources"]

    row_headers = {}
    col_headers = {}
    header_set = set()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val:
            row_headers[val.lower()] = r
            header_set.add(val.lower())
    for c in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            col_headers[val.lower()] = c

    release_summary = fetch_release_summary()
    sources = {k: v for k, v in release_summary.items() if k != 'translator_kg'}
    print(f"Processing {len(sources)} ingests from release summary")

    ctx = ssl.create_default_context()
    cell_data = defaultdict(lambda: defaultdict(set))

    for source_name, info in sorted(sources.items()):
        data_url = info['data'].rstrip('/') + '/'
        metadata_url = data_url + 'graph-metadata.json'

        try:
            with urllib.request.urlopen(metadata_url, context=ctx, timeout=30) as resp:
                metadata = json.loads(resp.read())
        except Exception as e:
            print(f"  FAILED {source_name}: {e}")
            continue

        edges = metadata.get('schema', {}).get('edges', [])
        if not edges:
            print(f"  {source_name}: no edges in schema")
            continue

        pairs_added = set()
        for edge in edges:
            subj_categories = edge.get('subject_category', [])
            obj_categories = edge.get('object_category', [])

            all_subj = set()
            for cat in subj_categories:
                name = curie_to_class_name(cat)
                all_subj.update(get_entity_ancestors(name, category_to_entity_classes))

            all_obj = set()
            for cat in obj_categories:
                name = curie_to_class_name(cat)
                all_obj.update(get_entity_ancestors(name, category_to_entity_classes))

            for s in all_subj:
                if s in row_headers:
                    for o in all_obj:
                        if o in col_headers:
                            cell_data[row_headers[s]][col_headers[o]].add(source_name)
                            pairs_added.add((s, o))

        print(f"  {source_name}: {len(edges)} edge groups -> {len(pairs_added)} matrix pairs")

    # Write to spreadsheet
    cells_written = 0
    for row_idx, cols in cell_data.items():
        for col_idx, ingests in cols.items():
            ws.cell(row=row_idx, column=col_idx).value = ", ".join(sorted(ingests))
            cells_written += 1

    print(f"Wrote {cells_written} cells")
    return header_set, row_headers, col_headers


def add_edge_comments(wb, local_data_path, category_to_entity_classes, header_set, row_headers, col_headers):
    """Scan local edge data and add example edge comments to cells."""
    ws = wb["Entity Matrix - Data Sources"]
    data_base = Path(local_data_path)

    if not data_base.exists():
        print(f"Local data path not found: {data_base}")
        return

    # Get filled cells and their ingests
    filled_cells = {}
    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val:
                filled_cells[(r, c)] = [x.strip() for x in val.split(',')]

    # Find ingests with local data
    ingest_data = {}
    for ingest_dir in sorted(data_base.iterdir()):
        if not ingest_dir.is_dir():
            continue
        name = ingest_dir.name
        edges_files = sorted(ingest_dir.rglob('normalized_edges.jsonl'))
        nodes_files = sorted(ingest_dir.rglob('normalized_nodes.jsonl'))
        if edges_files and nodes_files:
            ef = edges_files[-1]
            nf = nodes_files[-1]
            if ef.stat().st_size > 0 and nf.stat().st_size > 0:
                ingest_data[name] = {'edges': ef, 'nodes': nf}

    ingests_needed = set()
    for ingests in filled_cells.values():
        ingests_needed.update(ingests)

    cell_examples = defaultdict(dict)

    for pass_num in range(1, 3):
        scan_limit = MAX_EDGES_SCAN if pass_num == 1 else None
        pass_label = f"pass {pass_num}" + (" (full scan)" if pass_num == 2 else "")

        for ingest_name in sorted(ingests_needed):
            if ingest_name not in ingest_data:
                continue

            cells_needing = set()
            for (r, c), ingests in filled_cells.items():
                if ingest_name in ingests and ingest_name not in cell_examples.get((r, c), {}):
                    cells_needing.add((r, c))

            if not cells_needing:
                continue

            if pass_num == 2 and len(cells_needing) == 0:
                continue

            edges_path = ingest_data[ingest_name]['edges']
            nodes_path = ingest_data[ingest_name]['nodes']

            print(f"  {ingest_name} ({pass_label}): {len(cells_needing)} cells needed")

            # Load nodes
            node_info = {}
            with open(nodes_path) as f:
                for line in f:
                    node = json.loads(line)
                    nid = node['id']
                    categories = node.get('category', [])
                    entity_cls_set = set()
                    for cat in categories:
                        cls_name = curie_to_class_name(cat)
                        if cls_name in category_to_entity_classes:
                            entity_cls_set.update(category_to_entity_classes[cls_name])
                    entity_cls_set &= header_set
                    if entity_cls_set:
                        node_info[nid] = (entity_cls_set, node.get('name', nid))

            # Scan edges
            edges_scanned = 0
            with open(edges_path) as f:
                for line in f:
                    edge = json.loads(line)
                    subj_info = node_info.get(edge['subject'])
                    obj_info = node_info.get(edge['object'])

                    if not subj_info or not obj_info:
                        edges_scanned += 1
                        if scan_limit and edges_scanned >= scan_limit:
                            break
                        continue

                    subj_classes, subj_name = subj_info
                    obj_classes, obj_name = obj_info
                    predicate = edge.get('predicate', '?')

                    for sc in subj_classes:
                        if sc not in row_headers:
                            continue
                        r = row_headers[sc]
                        for oc in obj_classes:
                            if oc not in col_headers:
                                continue
                            c = col_headers[oc]
                            if (r, c) in cells_needing:
                                pred_short = predicate.replace('biolink:', '')
                                example = (f"{edge['subject']} ({subj_name}) "
                                           f"--{pred_short}--> "
                                           f"{edge['object']} ({obj_name})")
                                cell_examples[(r, c)][ingest_name] = example
                                cells_needing.discard((r, c))

                    edges_scanned += 1
                    if not cells_needing:
                        break
                    if scan_limit and edges_scanned >= scan_limit:
                        break

            remaining = len(cells_needing)
            print(f"    scanned {edges_scanned} edges, {remaining} cells still missing")
            del node_info

        # After pass 1, check if pass 2 is needed
        total_missing = sum(
            1 for (r, c), ingests in filled_cells.items()
            for i in ingests
            if i in ingest_data and i not in cell_examples.get((r, c), {})
        )
        if total_missing == 0:
            break
        if pass_num == 1:
            print(f"\n  {total_missing} examples still missing, starting full scan pass...")

    # Write comments
    comments_written = 0
    for (r, c), examples in cell_examples.items():
        if not examples:
            continue
        lines = [f"[{name}] {ex}" for name, ex in sorted(examples.items())]
        ws.cell(row=r, column=c).comment = Comment("\n".join(lines), "biolink-matrix")
        comments_written += 1

    print(f"Wrote {comments_written} comments")


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('--local-data', type=str, default=None,
                        help='Path to translator-ingests/data/ directory for edge examples')
    parser.add_argument('--schema', type=str, default='biolink-model.yaml',
                        help='Path to biolink-model.yaml schema file')
    parser.add_argument('--output', type=str, default='biolink_entity_matrix.xlsx',
                        help='Output xlsx path')
    args = parser.parse_args()

    print("Loading biolink model...")
    sv = SchemaView(args.schema)
    entity_classes, category_to_entity_classes = build_hierarchy(sv)
    print(f"Found {len(entity_classes)} entity classes")

    print("Creating blank matrix...")
    wb = create_blank_matrix(entity_classes, sv)

    print("Filling ingest coverage data...")
    header_set, row_headers, col_headers = fill_ingest_data(wb, category_to_entity_classes)

    if args.local_data:
        print("\nAdding example edge comments from local data...")
        add_edge_comments(wb, args.local_data, category_to_entity_classes,
                          header_set, row_headers, col_headers)

    wb.save(args.output)
    print(f"\nSaved to {args.output}")


if __name__ == '__main__':
    main()
